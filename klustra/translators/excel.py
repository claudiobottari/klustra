"""ExcelTranslator — one Excel file → N KnowledgeUnits (SPEC §4.2).

Supports .xlsx and .xlsm (openpyxl OOXML only).  .xls (BIFF) requires xlrd and
is a future extension.

Table detection uses structural anchors:
- Fully empty rows / columns separate distinct tables.
- A region ≥ 2 rows is a table; its first row(s) are the header.
- Multi-row header: row-1 has a merge spanning multiple columns AND row-2 is
  all-strings — headers are flattened as "Group/Sub".

Formulas: records carry {"value": <cached>, "formula": "=B2*C2"}.
Merged cells are exploded: every cell in the merge gets the top-left value.
Free text outside tables → kind="narrative" unit per sheet.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter

from klustra.core.knowledge_unit import KnowledgeUnit
from klustra.core.source_ref import SourceRef
from klustra.ingestion.translator import TranslateContext, TranslationResult, Translator

# ---------------------------------------------------------------------------
# Internal cell representation
# ---------------------------------------------------------------------------


@dataclass
class _Cell:
    value: Any
    formula: str | None = None
    is_merge_copy: bool = False


def _has_content(cell: _Cell | None) -> bool:
    return cell is not None and cell.value is not None and str(cell.value).strip() != ""


def _cell_str(cell: _Cell | None) -> str:
    """Escaped string for markdown table cells."""
    if not _has_content(cell):
        return ""
    v = cell.value  # type: ignore[union-attr]
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).replace("|", "\\|").replace("\n", " ").replace("\r", "")


def _record_value(cell: _Cell | None) -> Any:
    """Records representation: formula cells → {"value": …, "formula": "…"}."""
    if cell is None:
        return None
    if cell.formula is not None:
        return {"value": cell.value, "formula": cell.formula}
    return cell.value


# ---------------------------------------------------------------------------
# Grid building (merge expansion + formula detection)
# ---------------------------------------------------------------------------


def _build_grid(ws_f: Any, ws_v: Any) -> dict[tuple[int, int], _Cell]:
    """Return {(row, col): _Cell} with merged cells expanded and formulas noted.

    ws_f = workbook loaded with data_only=False  (gives formula strings)
    ws_v = workbook loaded with data_only=True   (gives cached computed values)
    Coordinates are 1-indexed to match openpyxl conventions.
    """
    if ws_f.min_row is None:
        return {}

    # Build redirect map: non-top-left merge positions → top-left position
    tl_map: dict[tuple[int, int], tuple[int, int]] = {}
    for mr in ws_f.merged_cells.ranges:
        tl = (mr.min_row, mr.min_col)
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                if (r, c) != tl:
                    tl_map[(r, c)] = tl

    grid: dict[tuple[int, int], _Cell] = {}

    for row in range(ws_f.min_row, ws_f.max_row + 1):
        for col in range(ws_f.min_column, ws_f.max_column + 1):
            src = tl_map.get((row, col), (row, col))
            is_copy = src != (row, col)

            cf = ws_f.cell(row=src[0], column=src[1])
            cv = ws_v.cell(row=src[0], column=src[1])

            formula: str | None = None
            if getattr(cf, "data_type", None) == "f":
                formula = str(cf.value)
                value = cv.value
            else:
                value = cf.value

            if value is not None or formula is not None:
                grid[(row, col)] = _Cell(value=value, formula=formula, is_merge_copy=is_copy)

    return grid


# ---------------------------------------------------------------------------
# Region detection
# ---------------------------------------------------------------------------


@dataclass
class _Region:
    min_row: int
    max_row: int
    min_col: int
    max_col: int

    @property
    def range_str(self) -> str:
        return (
            f"{get_column_letter(self.min_col)}{self.min_row}"
            f":{get_column_letter(self.max_col)}{self.max_row}"
        )

    @property
    def n_rows(self) -> int:
        return self.max_row - self.min_row + 1

    @property
    def n_cols(self) -> int:
        return self.max_col - self.min_col + 1


def _find_regions(
    grid: dict[tuple[int, int], _Cell],
    min_row: int,
    max_row: int,
    min_col: int,
    max_col: int,
) -> list[_Region]:
    """Split the used range into rectangular non-empty blobs.

    Blobs are separated first by fully empty rows, then by fully empty columns
    within each row band.
    """

    def row_empty(r: int) -> bool:
        return not any(_has_content(grid.get((r, c))) for c in range(min_col, max_col + 1))

    def col_empty_in(c: int, r0: int, r1: int) -> bool:
        return not any(_has_content(grid.get((r, c))) for r in range(r0, r1 + 1))

    # Identify row bands
    row_bands: list[tuple[int, int]] = []
    start: int | None = None
    for r in range(min_row, max_row + 1):
        if not row_empty(r):
            if start is None:
                start = r
        elif start is not None:
            row_bands.append((start, r - 1))
            start = None
    if start is not None:
        row_bands.append((start, max_row))

    regions: list[_Region] = []
    for r0, r1 in row_bands:
        col_bands: list[tuple[int, int]] = []
        cs: int | None = None
        for c in range(min_col, max_col + 1):
            if not col_empty_in(c, r0, r1):
                if cs is None:
                    cs = c
            elif cs is not None:
                col_bands.append((cs, c - 1))
                cs = None
        if cs is not None:
            col_bands.append((cs, max_col))

        for c0, c1 in col_bands:
            regions.append(_Region(r0, r1, c0, c1))

    return regions


# ---------------------------------------------------------------------------
# Header detection
# ---------------------------------------------------------------------------


def _row_looks_like_header(grid: dict[tuple[int, int], _Cell], row: int, region: _Region) -> bool:
    cells = [grid.get((row, c)) for c in range(region.min_col, region.max_col + 1)]
    non_empty: list[_Cell] = [c for c in cells if c is not None and _has_content(c)]
    return bool(non_empty) and all(isinstance(c.value, str) for c in non_empty)


def _row1_has_multicolumn_merge(ws_f: Any, region: _Region) -> bool:
    """True if any merged range in row-1 of region spans more than one column."""
    for mr in ws_f.merged_cells.ranges:
        if (
            mr.min_row == region.min_row
            and mr.max_col > mr.min_col
            and mr.min_col >= region.min_col
            and mr.max_col <= region.max_col
        ):
            return True
    return False


def _detect_header_rows(grid: dict[tuple[int, int], _Cell], region: _Region, ws_f: Any) -> int:
    """Return 1 or 2 header rows; 0 if the first row does not look like a header."""
    if region.n_rows < 2:
        return 0
    if not _row_looks_like_header(grid, region.min_row, region):
        return 0
    if (
        region.n_rows >= 3
        and _row1_has_multicolumn_merge(ws_f, region)
        and _row_looks_like_header(grid, region.min_row + 1, region)
    ):
        return 2
    return 1


# ---------------------------------------------------------------------------
# Column name generation
# ---------------------------------------------------------------------------


def _make_col_names(
    grid: dict[tuple[int, int], _Cell], region: _Region, n_header: int
) -> list[str]:
    """Generate column names, flattening 2-level headers as "Group/Sub"."""
    cols = range(region.min_col, region.max_col + 1)

    if n_header == 1:
        raw = [_cell_str(grid.get((region.min_row, c))) or f"col_{i}" for i, c in enumerate(cols)]
    else:
        raw = []
        for i, c in enumerate(cols):
            r1 = _cell_str(grid.get((region.min_row, c)))
            r2 = _cell_str(grid.get((region.min_row + 1, c)))
            if r1 and r2:
                raw.append(f"{r1}/{r2}")
            elif r1:
                raw.append(r1)
            elif r2:
                raw.append(r2)
            else:
                raw.append(f"col_{i}")

    # Deduplicate: first occurrence keeps the name, subsequent get _2, _3, …
    seen: dict[str, int] = {}
    result: list[str] = []
    for name in raw:
        if name in seen:
            seen[name] += 1
            result.append(f"{name}_{seen[name]}")
        else:
            seen[name] = 0
            result.append(name)
    return result


# ---------------------------------------------------------------------------
# Markdown table rendering
# ---------------------------------------------------------------------------


def _make_markdown_table(col_names: list[str], rows: list[list[_Cell | None]]) -> str:
    def esc(s: str) -> str:
        return s.replace("|", "\\|").replace("\n", " ")

    header = "| " + " | ".join(esc(n) for n in col_names) + " |"
    sep = "| " + " | ".join("---" for _ in col_names) + " |"
    body = ["| " + " | ".join(_cell_str(cell) for cell in row) + " |" for row in rows]
    return "\n".join([header, sep, *body])


# ---------------------------------------------------------------------------
# Unit extraction
# ---------------------------------------------------------------------------


def _title_above(grid: dict[tuple[int, int], _Cell], region: _Region) -> str | None:
    """Return the first non-empty string value in the row directly above the region."""
    if region.min_row <= 1:
        return None
    for c in range(region.min_col, region.max_col + 1):
        cell = grid.get((region.min_row - 1, c))
        if _has_content(cell) and isinstance(cell.value, str):  # type: ignore[union-attr]
            return cell.value  # type: ignore[union-attr]
    return None


def _extract_table_unit(
    grid: dict[tuple[int, int], _Cell],
    region: _Region,
    ws_f: Any,
    sheet_name: str,
    source_id: str,
    seq: int,
    file_props: dict[str, Any],
) -> KnowledgeUnit:
    n_header = _detect_header_rows(grid, region, ws_f)
    if n_header == 0:
        n_header = 1  # fallback: treat first row as header

    col_names = _make_col_names(grid, region, n_header)
    data_start = region.min_row + n_header

    data_rows: list[list[_Cell | None]] = [
        [grid.get((r, c)) for c in range(region.min_col, region.max_col + 1)]
        for r in range(data_start, region.max_row + 1)
    ]

    records: list[dict[str, Any]] = [
        {col: _record_value(cell) for col, cell in zip(col_names, row, strict=True)}
        for row in data_rows
    ]

    return KnowledgeUnit(
        unit_id=f"{source_id}#{seq}",
        kind="table",
        content_md=_make_markdown_table(col_names, data_rows),
        records=records,
        locator=f"sheet:{sheet_name}!{region.range_str}",
        inherited_context={
            "sheet_name": sheet_name,
            "title_row": _title_above(grid, region),
            "global_units": [],
            "file_props": file_props,
        },
    )


def _collect_narrative(
    grid: dict[tuple[int, int], _Cell],
    table_regions: list[_Region],
    sheet_name: str,
    source_id: str,
    seq: int,
    file_props: dict[str, Any],
) -> KnowledgeUnit | None:
    """Collect string cells outside every table region into a narrative unit."""
    in_table: set[tuple[int, int]] = {
        (r, c)
        for region in table_regions
        for r in range(region.min_row, region.max_row + 1)
        for c in range(region.min_col, region.max_col + 1)
    }

    parts: list[str] = []
    for (row, col), cell in sorted(grid.items()):
        if (row, col) not in in_table and not cell.is_merge_copy:
            if isinstance(cell.value, str) and cell.value.strip():
                parts.append(cell.value.strip())

    if not parts:
        return None

    return KnowledgeUnit(
        unit_id=f"{source_id}#{seq}",
        kind="narrative",
        content_md="\n\n".join(parts),
        locator=f"sheet:{sheet_name}",
        inherited_context={
            "sheet_name": sheet_name,
            "title_row": None,
            "global_units": [],
            "file_props": file_props,
        },
    )


# ---------------------------------------------------------------------------
# Sheet and file orchestration
# ---------------------------------------------------------------------------


def _get_file_props(wb: Any) -> dict[str, Any]:
    props: dict[str, Any] = {}
    if wb.properties:
        for attr in ("title", "creator", "description", "subject"):
            val = getattr(wb.properties, attr, None)
            if val:
                props[attr] = str(val)
    return props


def _translate_sheet(
    ws_f: Any,
    ws_v: Any,
    sheet_name: str,
    source: SourceRef,
    file_props: dict[str, Any],
    seq: int,
) -> tuple[list[KnowledgeUnit], list[str], int]:
    grid = _build_grid(ws_f, ws_v)
    if not grid:
        return [], [], seq

    rows = [r for r, _ in grid]
    cols = [c for _, c in grid]
    regions = _find_regions(grid, min(rows), max(rows), min(cols), max(cols))

    # Only regions with ≥2 rows become tables; single rows become narrative
    table_regions = [reg for reg in regions if reg.n_rows >= 2]

    units: list[KnowledgeUnit] = []
    for region in table_regions:
        units.append(
            _extract_table_unit(grid, region, ws_f, sheet_name, source.source_id, seq, file_props)
        )
        seq += 1

    narrative = _collect_narrative(
        grid, table_regions, sheet_name, source.source_id, seq, file_props
    )
    if narrative:
        units.append(narrative)
        seq += 1

    return units, [], seq


# ---------------------------------------------------------------------------
# Public translator
# ---------------------------------------------------------------------------


class ExcelTranslator(Translator):
    """Deterministic Excel → KnowledgeUnit translator (SPEC §4.2).

    Supports .xlsx and .xlsm (OOXML via openpyxl).
    .xls (BIFF) is not supported in v0.1 — add xlrd as a future extension.
    """

    name = "excel"
    version = "1.0"
    extensions = {".xlsx", ".xlsm"}

    def translate(self, source: SourceRef, ctx: TranslateContext) -> TranslationResult:
        path = Path(source.source_path)
        wb_f = openpyxl.load_workbook(path, data_only=False)
        wb_v = openpyxl.load_workbook(path, data_only=True)

        try:
            file_props = _get_file_props(wb_f)
            sheet_names = list(wb_f.sheetnames)
            units: list[KnowledgeUnit] = []
            warnings: list[str] = []
            seq = 0

            for name in sheet_names:
                sheet_units, sheet_warnings, seq = _translate_sheet(
                    wb_f[name], wb_v[name], name, source, file_props, seq
                )
                units.extend(sheet_units)
                warnings.extend(sheet_warnings)
        finally:
            wb_f.close()
            wb_v.close()

        return TranslationResult(
            units=units,
            source_metadata={"file_path": source.source_path, "sheets": sheet_names, **file_props},
            warnings=warnings,
        )
