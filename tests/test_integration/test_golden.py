"""Golden integration test: mini-corpus → compile → export okf_bundle → diff golden (SPEC §14)."""

from __future__ import annotations

import hashlib
import json
import shutil
from datetime import UTC, datetime
from pathlib import Path

import openpyxl
import pytest
from openpyxl.styles import Font

from klustra.api import Klustra
from klustra.llm.provider import LLMProvider, LLMRequest, LLMResponse

FIXTURES = Path(__file__).parent.parent / "fixtures"
MINI_CORPUS = FIXTURES / "mini_corpus"
GOLDEN_DIR = FIXTURES / "golden"

FIXED_TIME = datetime(2025, 1, 15, 12, 0, 0, tzinfo=UTC)


def _deterministic_source_id(path: Path) -> str:
    """Source ID based on filename only — portable across machines."""
    return hashlib.sha256(path.name.encode()).hexdigest()[:16]


def _build_project_report(path: Path) -> None:
    """Generate a messy Excel fixture: budget table + timeline table."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Budget"

    ws.merge_cells("A1:D1")
    ws["A1"] = "HV Cable Project Budget 2024"
    ws["A1"].font = Font(bold=True)

    for c, h in enumerate(["Phase", "Budget (EUR)", "Spent", "Remaining"], start=1):
        ws.cell(row=3, column=c, value=h)
    for r, (phase, budget, spent) in enumerate(
        [
            ("Design", 250_000, 245_000),
            ("Procurement", 1_200_000, 980_000),
            ("Installation", 800_000, 320_000),
            ("Testing", 150_000, 45_000),
        ],
        start=4,
    ):
        ws.cell(row=r, column=1, value=phase)
        ws.cell(row=r, column=2, value=budget)
        ws.cell(row=r, column=3, value=spent)
        ws.cell(row=r, column=4, value=f"=B{r}-C{r}")

    ws2 = wb.create_sheet("Timeline")
    for c, h in enumerate(["Milestone", "Planned", "Actual", "Status"], start=1):
        ws2.cell(row=1, column=c, value=h)
    for r, row_vals in enumerate(
        [
            ("Route survey", "2024-01", "2024-01", "Complete"),
            ("Cable delivery", "2024-03", "2024-04", "Delayed"),
            ("Installation start", "2024-05", "2024-06", "Delayed"),
            ("Commissioning", "2024-09", "", "Pending"),
        ],
        start=2,
    ):
        for c, val in enumerate(row_vals, start=1):
            ws2.cell(row=r, column=c, value=val)

    wb.save(path)


# ---------------------------------------------------------------------------
# Mock provider — deterministic responses keyed by content
# ---------------------------------------------------------------------------

# Source IDs for citation references (deterministic, based on filename)
_SRC_IDS = {
    name: hashlib.sha256(name.encode()).hexdigest()[:16]
    for name in [
        "cable_intro.md",
        "xlpe_material.md",
        "testing_procedures.md",
        "installation_guide.txt",
        "safety_standards.md",
        "supplier_list.txt",
        "project_report.xlsx",
    ]
}


def _extraction_response_for(user_content: str) -> dict:
    """Route extraction by document title/heading in the unit content."""
    lower = user_content.lower()

    # Route by document heading (most specific first)
    if "xlpe insulation material" in lower:
        return {
            "candidates": [
                {
                    "name": "XLPE Insulation",
                    "entity_id_proposal": "mat.xlpe",
                    "summary": "Cross-linked polyethylene insulation material.",
                    "is_new": True,
                    "related_existing": [],
                }
            ]
        }
    if "cable testing standards" in lower:
        return {
            "candidates": [
                {
                    "name": "Cable Testing",
                    "entity_id_proposal": "proc.testing",
                    "summary": "Testing procedures for HV cable systems.",
                    "is_new": True,
                    "related_existing": [],
                }
            ]
        }
    if "cable installation guide" in lower:
        return {
            "candidates": [
                {
                    "name": "Cable Installation",
                    "entity_id_proposal": "proc.installation",
                    "summary": "Installation procedures for HV cables.",
                    "is_new": True,
                    "related_existing": [],
                }
            ]
        }
    # Everything else contributes to the main HV cable entity
    return {
        "candidates": [
            {
                "name": "HV Cable System",
                "entity_id_proposal": "prod.hv-cable",
                "summary": "High-voltage cable system for power transmission.",
                "is_new": True,
                "related_existing": [],
            }
        ]
    }


_SID_CABLE = _SRC_IDS["cable_intro.md"]
_SID_XLPE = _SRC_IDS["xlpe_material.md"]
_SID_TEST = _SRC_IDS["testing_procedures.md"]
_SID_INSTALL = _SRC_IDS["installation_guide.txt"]
_SID_SAFETY = _SRC_IDS["safety_standards.md"]
_SID_SUPPLIER = _SRC_IDS["supplier_list.txt"]
_SID_REPORT = _SRC_IDS["project_report.xlsx"]


_LIBRARIAN_RESPONSES: dict[str, dict] = {
    "prod.hv-cable": {
        "title": "HV Cable System",
        "description": "High-voltage cable system for underground power transmission.",
        "body_md": (
            "High-voltage cable systems are critical infrastructure "
            "for power transmission, "
            f"operating at voltages up to 500 kV. ^[{_SID_CABLE}:doc:1]\n\n"
            "The cable core uses a copper or aluminium conductor "
            f"with [[mat.xlpe]] insulation. ^[{_SID_CABLE}:doc:1]\n\n"
            "## Project Status\n\n"
            "Current project budget is EUR 2.4M across four phases. "
            f"^[{_SID_REPORT}:sheet:Budget!A3:D7]\n\n"
            "## Safety\n\n"
            "All work must comply with IEC 61936-1 and local "
            f"electrical safety regulations. ^[{_SID_SAFETY}:doc:1]"
        ),
        "tags": ["cable", "hv", "power-transmission"],
        "aliases": ["HV Cable", "High-Voltage Cable"],
        "confidence": 0.92,
    },
    "mat.xlpe": {
        "title": "XLPE Insulation",
        "description": "Cross-linked polyethylene insulation for high-voltage cables.",
        "body_md": (
            "Cross-linked polyethylene (XLPE) is the dominant insulation "
            "material for [[prod.hv-cable]] systems operating at 66 kV "
            f"and above. ^[{_SID_XLPE}:doc:1]\n\n"
            "## Properties\n\n"
            "XLPE provides dielectric strength of 20-30 kV/mm with "
            f"operating temperature up to 90°C. ^[{_SID_XLPE}:doc:1]\n\n"
            "## Suppliers\n\n"
            "Approved compounds: Borealis Visico LE4253 and "
            f"Dow HFDA-4202 EC. ^[{_SID_SUPPLIER}:doc:1]"
        ),
        "tags": ["material", "insulation", "xlpe"],
        "aliases": ["Cross-linked Polyethylene", "XLPE"],
        "confidence": 0.95,
    },
    "proc.testing": {
        "title": "Cable Testing",
        "description": "Testing procedures and standards for HV cable systems.",
        "body_md": (
            "Cable testing follows IEC 62067 and covers type tests, "
            "routine tests, and after-installation tests. "
            f"^[{_SID_TEST}:doc:1]\n\n"
            "## Type Testing\n\n"
            "Validates cable system design through electrical, "
            f"mechanical, and thermal tests. ^[{_SID_TEST}:doc:1]\n\n"
            "## Routine Testing\n\n"
            "Every manufactured length undergoes HV withstand "
            f"and PD measurement. ^[{_SID_TEST}:doc:1]"
        ),
        "tags": ["testing", "standards", "iec"],
        "aliases": ["HV Cable Testing"],
        "confidence": 0.90,
    },
    "proc.installation": {
        "title": "Cable Installation",
        "description": "Installation procedures and requirements for HV cables.",
        "body_md": (
            "Cable installation requires minimum bending radius of "
            "20x outer diameter and maximum pulling tension of "
            f"50 N/mm2. ^[{_SID_INSTALL}:doc:1]\n\n"
            "## Methods\n\n"
            "Installation methods include duct pulling, direct burial, "
            f"and trenchless techniques. ^[{_SID_INSTALL}:doc:1]\n\n"
            "## Jointing\n\n"
            "All joints must be installed in a controlled "
            "environment with prefabricated joints preferred. "
            f"^[{_SID_INSTALL}:doc:1]"
        ),
        "tags": ["installation", "construction"],
        "aliases": ["HV Cable Installation"],
        "confidence": 0.88,
    },
}


class _GoldenMockProvider(LLMProvider):
    """Deterministic provider for golden tests — routes by content keywords."""

    name = "golden_mock"

    def call(self, request: LLMRequest) -> LLMResponse:
        system_content = request.messages[0].content if request.messages else ""
        user_content = request.messages[1].content if len(request.messages) > 1 else ""

        if "extraction" in system_content.lower():
            data = _extraction_response_for(user_content)
        else:
            data = self._librarian_response(user_content)

        content = json.dumps(data)
        msg_chars = sum(len(m.content) for m in request.messages)
        return LLMResponse(
            content=content,
            parsed=data,
            tokens_in=max(1, msg_chars // 4),
            tokens_out=max(1, len(content) // 4),
            model=request.model,
        )

    def _librarian_response(self, user_content: str) -> dict:
        for entity_id, response in _LIBRARIAN_RESPONSES.items():
            if f"## Entity: {entity_id}" in user_content:
                return response
        return _LIBRARIAN_RESPONSES["prod.hv-cable"]


# ---------------------------------------------------------------------------
# Normalization — make output portable across machines
# ---------------------------------------------------------------------------


def _normalize_file(content: str, corpus_dir: str) -> str:
    """Replace machine-specific paths and timestamps with canonical values."""
    import re

    normalized = corpus_dir.replace("\\", "/")
    content = content.replace("\\", "/")
    content = content.replace(normalized, "CORPUS/")
    # Normalize ISO timestamps to fixed value
    content = re.sub(
        r"'\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}\.\d+\+00:00'",
        "'2025-01-15T12:00:00.000000+00:00'",
        content,
    )
    return content


def _normalize_output(output_dir: Path, corpus_dir: Path) -> dict[str, str]:
    """Read all files in output_dir and return normalized {relative_path: content}."""
    corpus_str = str(corpus_dir).rstrip("/\\") + "/"
    result: dict[str, str] = {}
    for f in sorted(output_dir.rglob("*")):
        if f.is_file():
            rel = f.relative_to(output_dir).as_posix()
            content = f.read_text(encoding="utf-8")
            result[rel] = _normalize_file(content, corpus_str)
    return result


def _read_golden() -> dict[str, str]:
    """Read golden bundle from fixtures."""
    if not GOLDEN_DIR.exists():
        return {}
    result: dict[str, str] = {}
    for f in sorted(GOLDEN_DIR.rglob("*")):
        if f.is_file():
            rel = f.relative_to(GOLDEN_DIR).as_posix()
            result[rel] = f.read_text(encoding="utf-8")
    return result


def _write_golden(files: dict[str, str]) -> None:
    """Write normalized output as the new golden."""
    if GOLDEN_DIR.exists():
        shutil.rmtree(GOLDEN_DIR)
    for rel_path, content in files.items():
        out = GOLDEN_DIR / rel_path
        out.parent.mkdir(parents=True, exist_ok=True)
        out.write_text(content, encoding="utf-8")


# ---------------------------------------------------------------------------
# Test
# ---------------------------------------------------------------------------


@pytest.fixture
def golden_project(tmp_path: Path) -> Path:
    """Set up a klustra project with mini-corpus (including generated Excel)."""
    root = tmp_path / "project"
    root.mkdir()

    (root / "klustra.toml").write_text(
        '[llm.extraction]\nprovider = "mock"\nmodel = "golden-model"\n\n'
        '[llm.librarian]\nprovider = "mock"\nmodel = "golden-model"\n',
        encoding="utf-8",
    )

    for subdir in [".klustra", ".klustra/domains", ".klustra/instructions", ".klustra/vault"]:
        (root / subdir).mkdir(parents=True, exist_ok=True)

    corpus_dir = root / "corpus"
    corpus_dir.mkdir()

    for f in MINI_CORPUS.iterdir():
        if f.is_file():
            shutil.copy2(f, corpus_dir / f.name)

    _build_project_report(corpus_dir / "project_report.xlsx")

    return root


def test_golden_bundle(
    golden_project: Path,
    update_goldens: bool,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """Full pipeline golden test: ingest → compile → export → diff golden."""
    # Monkeypatch for determinism
    monkeypatch.setattr("klustra.ingestion.source_manager._source_id", _deterministic_source_id)

    _fixed_now = FIXED_TIME

    class _FrozenDatetime(datetime):
        @classmethod
        def now(cls, tz=None):  # noqa: ANN001, ANN206
            return _fixed_now

    monkeypatch.setattr("klustra.engine.librarian.datetime", _FrozenDatetime)

    _uuid_counter = iter(range(1000))

    class _FakeUUID:
        def __init__(self, val: int) -> None:
            self.hex = f"{val:032x}"

        def __str__(self) -> str:
            return f"00000000-0000-0000-0000-{next(_uuid_counter):012d}"

    def _fake_uuid4():  # noqa: ANN202
        return _FakeUUID(0)

    monkeypatch.setattr("klustra.api.uuid.uuid4", _fake_uuid4)
    monkeypatch.setattr("klustra.ingestion.source_manager.uuid.uuid4", _fake_uuid4)

    # Run pipeline
    corpus_dir = golden_project / "corpus"
    provider = _GoldenMockProvider()
    nx = Klustra(root=golden_project, provider=provider)

    nx.ingest_folder(corpus_dir)
    results = nx.compile()
    assert len(results) >= 1

    # Write body_md to vault (compile doesn't do this yet)
    for result in results:
        vault_path = golden_project / ".klustra" / "vault" / f"{result.page.entity_id}.md"
        vault_path.parent.mkdir(parents=True, exist_ok=True)
        vault_path.write_text(result.body_md, encoding="utf-8")

    # Export OKF bundle
    export_dir = golden_project / "export"
    nx.export("okf_bundle", export_dir)

    # Normalize output
    actual = _normalize_output(export_dir, corpus_dir)
    assert actual, "Export produced no files"

    if update_goldens:
        _write_golden(actual)
        pytest.skip("Golden files updated — re-run without --update-goldens to verify.")

    # Compare against golden
    golden = _read_golden()
    assert golden, (
        f"Golden files not found at {GOLDEN_DIR}. Run with --update-goldens to generate them."
    )

    # Check file sets match
    actual_keys = set(actual.keys())
    golden_keys = set(golden.keys())
    missing = golden_keys - actual_keys
    extra = actual_keys - golden_keys
    assert not missing, f"Files missing from output (present in golden): {missing}"
    assert not extra, f"Extra files in output (not in golden): {extra}"

    # Compare content
    diffs: list[str] = []
    for path in sorted(actual_keys):
        if actual[path] != golden[path]:
            diffs.append(path)

    if diffs:
        # Show first diff for debugging
        first = diffs[0]
        msg_lines = [f"Golden mismatch in {len(diffs)} file(s). First: {first}"]
        msg_lines.append("--- GOLDEN ---")
        msg_lines.append(golden[first][:500])
        msg_lines.append("--- ACTUAL ---")
        msg_lines.append(actual[first][:500])
        pytest.fail("\n".join(msg_lines))
