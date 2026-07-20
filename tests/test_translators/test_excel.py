"""Tests for ExcelTranslator (SPEC §4.2)."""

from pathlib import Path
from typing import Any

import openpyxl

from klustra.translators.excel import (
    ExcelTranslator,
    _build_grid,
    _Cell,
    _detect_header_rows,
    _find_regions,
    _make_col_names,
    _Region,
)
from klustra.translators.registry import build_default_registry
from tests.test_translators.conftest import make_ctx, make_source

TR = ExcelTranslator()
CTX = make_ctx()


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def units_by_kind(result: Any, kind: str) -> list[Any]:
    return [u for u in result.units if u.kind == kind]


# ---------------------------------------------------------------------------
# _build_grid unit tests
# ---------------------------------------------------------------------------


def test_build_grid_basic(tmp_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "hello"
    ws["B2"] = 42
    tmp = tmp_path / "t.xlsx"
    wb.save(tmp)
    wb_f = openpyxl.load_workbook(tmp, data_only=False)
    wb_v = openpyxl.load_workbook(tmp, data_only=True)
    grid = _build_grid(wb_f.active, wb_v.active)
    assert grid[(1, 1)].value == "hello"
    assert grid[(2, 2)].value == 42


def test_build_grid_merge_expanded(tmp_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.merge_cells("A1:C1")
    ws["A1"] = "Merged"
    tmp = tmp_path / "t.xlsx"
    wb.save(tmp)
    wb_f = openpyxl.load_workbook(tmp, data_only=False)
    wb_v = openpyxl.load_workbook(tmp, data_only=True)
    grid = _build_grid(wb_f.active, wb_v.active)
    # All three columns get the merged value
    assert grid[(1, 1)].value == "Merged"
    assert grid[(1, 2)].value == "Merged"
    assert grid[(1, 3)].value == "Merged"
    # Non-TL cells are flagged as copies
    assert not grid[(1, 1)].is_merge_copy
    assert grid[(1, 2)].is_merge_copy
    assert grid[(1, 3)].is_merge_copy


def test_build_grid_formula_captured(tmp_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = 10
    ws["B1"] = "=A1*3"
    tmp = tmp_path / "t.xlsx"
    wb.save(tmp)
    wb_f = openpyxl.load_workbook(tmp, data_only=False)
    wb_v = openpyxl.load_workbook(tmp, data_only=True)
    grid = _build_grid(wb_f.active, wb_v.active)
    b1 = grid[(1, 2)]
    assert b1.formula == "=A1*3"
    # value is None for synthetic files (no Excel to cache it)
    assert b1.value is None


# ---------------------------------------------------------------------------
# _find_regions unit tests
# ---------------------------------------------------------------------------


def test_find_regions_single_blob() -> None:
    grid = {
        (1, 1): _Cell("h1"),
        (1, 2): _Cell("h2"),
        (2, 1): _Cell("d1"),
        (2, 2): _Cell("d2"),
    }
    regions = _find_regions(grid, 1, 2, 1, 2)
    assert len(regions) == 1
    assert regions[0] == _Region(1, 2, 1, 2)


def test_find_regions_two_tables_separated_by_empty_row() -> None:
    grid = {
        (1, 1): _Cell("H1"),
        (1, 2): _Cell("H2"),
        (2, 1): _Cell("d1"),
        (2, 2): _Cell("d2"),
        # row 3 empty
        (4, 1): _Cell("H3"),
        (4, 2): _Cell("H4"),
        (5, 1): _Cell("d3"),
        (5, 2): _Cell("d4"),
    }
    regions = _find_regions(grid, 1, 5, 1, 2)
    assert len(regions) == 2
    assert regions[0] == _Region(1, 2, 1, 2)
    assert regions[1] == _Region(4, 5, 1, 2)


def test_find_regions_two_tables_side_by_side() -> None:
    grid = {
        (1, 1): _Cell("A"),
        (2, 1): _Cell("a1"),
        (1, 3): _Cell("B"),
        (2, 3): _Cell("b1"),
    }
    regions = _find_regions(grid, 1, 2, 1, 3)
    assert len(regions) == 2


def test_find_regions_empty_sheet() -> None:
    assert _find_regions({}, 1, 5, 1, 5) == []


# ---------------------------------------------------------------------------
# _detect_header_rows unit tests
# ---------------------------------------------------------------------------


def test_detect_header_single_row(tmp_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Name"
    ws["B1"] = "Value"
    ws["A2"] = "Alice"
    ws["B2"] = 42
    tmp = tmp_path / "t.xlsx"
    wb.save(tmp)
    wb_f = openpyxl.load_workbook(tmp, data_only=False)
    wb_v = openpyxl.load_workbook(tmp, data_only=True)
    grid = _build_grid(wb_f.active, wb_v.active)
    region = _Region(1, 2, 1, 2)
    assert _detect_header_rows(grid, region, wb_f.active) == 1


def test_detect_header_two_rows_merged(tmp_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.merge_cells("A1:C1")
    ws["A1"] = "Revenue"
    ws["A2"] = "Q1"
    ws["B2"] = "Q2"
    ws["C2"] = "Q3"
    for r in range(3, 6):
        for c in range(1, 4):
            ws.cell(row=r, column=c, value=r * c * 100)
    tmp = tmp_path / "t.xlsx"
    wb.save(tmp)
    wb_f = openpyxl.load_workbook(tmp, data_only=False)
    wb_v = openpyxl.load_workbook(tmp, data_only=True)
    grid = _build_grid(wb_f.active, wb_v.active)
    region = _Region(1, 5, 1, 3)
    assert _detect_header_rows(grid, region, wb_f.active) == 2


# ---------------------------------------------------------------------------
# _make_col_names unit tests
# ---------------------------------------------------------------------------


def test_make_col_names_single_header() -> None:
    grid = {
        (1, 1): _Cell("Name"),
        (1, 2): _Cell("Age"),
        (2, 1): _Cell("Alice"),
        (2, 2): _Cell(30),
    }
    names = _make_col_names(grid, _Region(1, 2, 1, 2), 1)
    assert names == ["Name", "Age"]


def test_make_col_names_deduplication() -> None:
    grid = {(1, c): _Cell("Score") for c in range(1, 4)}
    grid.update({(2, c): _Cell(c * 10) for c in range(1, 4)})
    names = _make_col_names(grid, _Region(1, 2, 1, 3), 1)
    assert names == ["Score", "Score_1", "Score_2"]


def test_make_col_names_two_level() -> None:
    grid = {
        (1, 1): _Cell("Revenue"),
        (1, 2): _Cell("Revenue"),
        (1, 3): _Cell("Costs"),
        (2, 1): _Cell("Q1"),
        (2, 2): _Cell("Q2"),
        (2, 3): _Cell("Q1"),
        (3, 1): _Cell(100),
    }
    names = _make_col_names(grid, _Region(1, 3, 1, 3), 2)
    assert names[0] == "Revenue/Q1"
    assert names[1] == "Revenue/Q2"
    assert names[2] == "Costs/Q1"


# ---------------------------------------------------------------------------
# ExcelTranslator.translate — messy multi-table fixture
# ---------------------------------------------------------------------------


def test_messy_two_table_units(xlsx_messy: Path) -> None:
    result = TR.translate(make_source(xlsx_messy), CTX)
    tables = units_by_kind(result, "table")
    assert len(tables) == 2


def test_messy_table_locators(xlsx_messy: Path) -> None:
    result = TR.translate(make_source(xlsx_messy), CTX)
    locators = {u.locator for u in result.units if u.kind == "table"}
    assert all(loc.startswith("sheet:Data!") for loc in locators)


def test_messy_employee_table_columns(xlsx_messy: Path) -> None:
    result = TR.translate(make_source(xlsx_messy), CTX)
    tables = units_by_kind(result, "table")
    # First table has Name, Dept, Salary, Grade columns
    emp_table = tables[0]
    assert emp_table.records is not None
    assert len(emp_table.records) == 4
    assert set(emp_table.records[0].keys()) == {"Name", "Dept", "Salary", "Grade"}


def test_messy_records_typed(xlsx_messy: Path) -> None:
    result = TR.translate(make_source(xlsx_messy), CTX)
    tables = units_by_kind(result, "table")
    emp_table = tables[0]
    assert emp_table.records is not None
    salaries = [r["Salary"] for r in emp_table.records]
    assert all(isinstance(s, int) for s in salaries)


def test_messy_formula_in_records(xlsx_messy: Path) -> None:
    result = TR.translate(make_source(xlsx_messy), CTX)
    tables = units_by_kind(result, "table")
    proj_table = tables[1]
    assert proj_table.records is not None
    remaining_vals = [r["Remaining"] for r in proj_table.records]
    # All Remaining cells have a formula
    assert all(isinstance(v, dict) and "formula" in v for v in remaining_vals)
    assert all("=B" in v["formula"] for v in remaining_vals)


def test_messy_content_md_is_markdown(xlsx_messy: Path) -> None:
    result = TR.translate(make_source(xlsx_messy), CTX)
    for unit in units_by_kind(result, "table"):
        assert unit.content_md.startswith("|")
        assert "---" in unit.content_md


def test_messy_narrative_unit(xlsx_messy: Path) -> None:
    result = TR.translate(make_source(xlsx_messy), CTX)
    narratives = units_by_kind(result, "narrative")
    assert len(narratives) == 1
    assert "Quarterly Report 2024" in narratives[0].content_md
    assert "Confidential" in narratives[0].content_md


def test_messy_narrative_locator(xlsx_messy: Path) -> None:
    result = TR.translate(make_source(xlsx_messy), CTX)
    narratives = units_by_kind(result, "narrative")
    assert narratives[0].locator == "sheet:Data"


def test_messy_inherited_context_keys(xlsx_messy: Path) -> None:
    result = TR.translate(make_source(xlsx_messy), CTX)
    for unit in result.units:
        ctx = unit.inherited_context
        assert "sheet_name" in ctx
        assert "title_row" in ctx
        assert "global_units" in ctx
        assert "file_props" in ctx


def test_messy_unit_ids_deterministic(xlsx_messy: Path) -> None:
    src = make_source(xlsx_messy, source_id="src-xls")
    r1 = TR.translate(src, CTX)
    r2 = TR.translate(src, CTX)
    assert [u.unit_id for u in r1.units] == [u.unit_id for u in r2.units]


def test_messy_source_metadata(xlsx_messy: Path) -> None:
    result = TR.translate(make_source(xlsx_messy), CTX)
    assert result.source_metadata["sheets"] == ["Data"]
    assert "file_path" in result.source_metadata


# ---------------------------------------------------------------------------
# ExcelTranslator.translate — merged headers fixture
# ---------------------------------------------------------------------------


def test_merged_headers_two_header_rows(xlsx_merged_headers: Path) -> None:
    result = TR.translate(make_source(xlsx_merged_headers), CTX)
    tables = units_by_kind(result, "table")
    assert len(tables) == 1
    assert tables[0].records is not None
    col_names = list(tables[0].records[0].keys())
    # Should have flattened names like "Revenue/Q1"
    assert any("/" in name for name in col_names)


def test_merged_headers_column_name_format(xlsx_merged_headers: Path) -> None:
    result = TR.translate(make_source(xlsx_merged_headers), CTX)
    tables = units_by_kind(result, "table")
    col_names = list(tables[0].records[0].keys())  # type: ignore[index]
    assert "Revenue/Q1" in col_names
    assert "Costs/Q1" in col_names or "Costs/Q1_1" in col_names


def test_merged_headers_data_rows(xlsx_merged_headers: Path) -> None:
    result = TR.translate(make_source(xlsx_merged_headers), CTX)
    tables = units_by_kind(result, "table")
    assert len(tables[0].records) == 3  # type: ignore[arg-type]


def test_merged_headers_all_numeric(xlsx_merged_headers: Path) -> None:
    result = TR.translate(make_source(xlsx_merged_headers), CTX)
    tables = units_by_kind(result, "table")
    for record in tables[0].records:  # type: ignore[union-attr]
        for val in record.values():
            assert isinstance(val, (int, float))


# ---------------------------------------------------------------------------
# ExcelTranslator.translate — multi-sheet fixture
# ---------------------------------------------------------------------------


def test_multi_sheet_units_from_both_sheets(xlsx_multi_sheet: Path) -> None:
    result = TR.translate(make_source(xlsx_multi_sheet), CTX)
    sheet_names = {u.locator.split("!")[0].replace("sheet:", "") for u in result.units}
    assert "Inventory" in sheet_names
    assert "Summary" in sheet_names


def test_multi_sheet_source_metadata(xlsx_multi_sheet: Path) -> None:
    result = TR.translate(make_source(xlsx_multi_sheet), CTX)
    assert set(result.source_metadata["sheets"]) == {"Inventory", "Summary"}


def test_multi_sheet_total_table_count(xlsx_multi_sheet: Path) -> None:
    result = TR.translate(make_source(xlsx_multi_sheet), CTX)
    tables = units_by_kind(result, "table")
    assert len(tables) >= 2  # at least one table per sheet


def test_multi_sheet_summary_narrative(xlsx_multi_sheet: Path) -> None:
    result = TR.translate(make_source(xlsx_multi_sheet), CTX)
    narratives = units_by_kind(result, "narrative")
    combined = " ".join(u.content_md for u in narratives)
    assert "Executive Summary" in combined


# ---------------------------------------------------------------------------
# Edge cases
# ---------------------------------------------------------------------------


def test_empty_sheet_no_units(tmp_path: Path) -> None:
    wb = openpyxl.Workbook()
    p = tmp_path / "empty.xlsx"
    wb.save(p)
    result = TR.translate(make_source(p), CTX)
    assert result.units == []


def test_single_row_region_no_table(tmp_path: Path) -> None:
    """A single-row region should not produce a table unit."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Just one row of data here"
    p = tmp_path / "single.xlsx"
    wb.save(p)
    result = TR.translate(make_source(p), CTX)
    tables = units_by_kind(result, "table")
    assert len(tables) == 0


def test_pipe_in_value_escaped(tmp_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Name"
    ws["B1"] = "Value"
    ws["A2"] = "a|b"
    ws["B2"] = 1
    p = tmp_path / "pipe.xlsx"
    wb.save(p)
    result = TR.translate(make_source(p), CTX)
    tables = units_by_kind(result, "table")
    assert len(tables) == 1
    assert "\\|" in tables[0].content_md


def test_all_units_have_kind_table_or_narrative(xlsx_messy: Path) -> None:
    result = TR.translate(make_source(xlsx_messy), CTX)
    for unit in result.units:
        assert unit.kind in ("table", "narrative")


# ---------------------------------------------------------------------------
# Registry
# ---------------------------------------------------------------------------


def test_registry_has_xlsx() -> None:
    reg = build_default_registry()
    assert reg.get_for_path(Path("report.xlsx")).name == "excel"


def test_registry_has_xlsm() -> None:
    reg = build_default_registry()
    assert reg.get_for_path(Path("report.xlsm")).name == "excel"


def test_extensions() -> None:
    assert ".xlsx" in TR.extensions
    assert ".xlsm" in TR.extensions
