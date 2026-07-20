"""Shared fixtures for test_translators."""

from pathlib import Path

import openpyxl
import pytest
from openpyxl.styles import Font

from klustra.core.source_ref import SourceRef
from klustra.ingestion.translator import TranslateContext

FIXTURES = Path(__file__).parent.parent / "fixtures" / "translators"


def make_source(path: Path, source_id: str = "test0001") -> SourceRef:
    return SourceRef(source_id=source_id, source_path=str(path))


def make_ctx(run_id: str = "run-test") -> TranslateContext:
    return TranslateContext(run_id=run_id)


@pytest.fixture
def multi_section_md() -> Path:
    return FIXTURES / "multi_section.md"


@pytest.fixture
def no_heading_md() -> Path:
    return FIXTURES / "no_heading.md"


@pytest.fixture
def plain_txt() -> Path:
    return FIXTURES / "plain.txt"


# ---------------------------------------------------------------------------
# Excel fixtures — generated programmatically (no binary blobs in git)
# ---------------------------------------------------------------------------


def _build_messy_multi_table(path: Path) -> None:
    """One sheet: merged title, two tables (second has a formula column), footnote.

    Rows (1-indexed):
      1     "Quarterly Report 2024"  [merged A1:D1]  — narrative
      2     (empty)
      3-7   Employee table: Name | Dept | Salary | Grade
      8     (empty)
      9-12  Project table: Project | Budget | Spent | Remaining  (=B-C formula)
      13    (empty)
      14    "Confidential — Internal Use Only"                   — narrative
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    ws.merge_cells("A1:D1")
    ws["A1"] = "Quarterly Report 2024"
    ws["A1"].font = Font(bold=True)

    for c, h in enumerate(["Name", "Dept", "Salary", "Grade"], start=1):
        ws.cell(row=3, column=c, value=h)
    for r, (name, dept, salary, grade) in enumerate(
        [
            ("Alice Bianchi", "Engineering", 72_000, "Senior"),
            ("Bob Müller", "Finance", 61_000, "Mid"),
            ("Carol López", "Engineering", 85_000, "Lead"),
            ("David Kim", "HR", 55_000, "Junior"),
        ],
        start=4,
    ):
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=dept)
        ws.cell(row=r, column=3, value=salary)
        ws.cell(row=r, column=4, value=grade)

    for c, h in enumerate(["Project", "Budget", "Spent", "Remaining"], start=1):
        ws.cell(row=9, column=c, value=h)
    for r, (proj, budget, spent) in enumerate(
        [("Alpha", 120_000, 45_000), ("Beta", 200_000, 198_500), ("Gamma", 50_000, 12_300)],
        start=10,
    ):
        ws.cell(row=r, column=1, value=proj)
        ws.cell(row=r, column=2, value=budget)
        ws.cell(row=r, column=3, value=spent)
        ws.cell(row=r, column=4, value=f"=B{r}-C{r}")  # formula

    ws["A14"] = "Confidential — Internal Use Only"
    wb.save(path)


def _build_merged_headers(path: Path) -> None:
    """One sheet: 2-level merged column headers.

    Row 1: "Revenue" [A1:C1]  |  "Costs" [D1:F1]
    Row 2: Q1 | Q2 | Q3       |  Q1 | Q2 | Q3
    Rows 3-5: numeric data
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Financials"

    ws.merge_cells("A1:C1")
    ws["A1"] = "Revenue"
    ws.merge_cells("D1:F1")
    ws["D1"] = "Costs"

    for c, label in enumerate(["Q1", "Q2", "Q3", "Q1", "Q2", "Q3"], start=1):
        ws.cell(row=2, column=c, value=label)

    for r, row_vals in enumerate(
        [
            [10_000, 12_500, 11_200, 8_000, 9_100, 7_800],
            [14_300, 13_800, 15_200, 9_500, 10_200, 11_000],
            [11_100, 10_500, 9_800, 8_700, 7_900, 8_200],
        ],
        start=3,
    ):
        for c, val in enumerate(row_vals, start=1):
            ws.cell(row=r, column=c, value=val)

    wb.save(path)


def _build_multi_sheet(path: Path) -> None:
    """Two sheets: Inventory (simple table) + Summary (narrative + KPI table)."""
    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "Inventory"
    for c, h in enumerate(["SKU", "Description", "Qty", "Unit Price"], start=1):
        ws1.cell(row=1, column=c, value=h)
    for r, row_vals in enumerate(
        [
            ("A001", "Widget A", 150, 9.99),
            ("A002", "Widget B", 240, 14.50),
            ("B001", "Gadget X", 30, 99.00),
        ],
        start=2,
    ):
        for c, val in enumerate(row_vals, start=1):
            ws1.cell(row=r, column=c, value=val)

    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "Executive Summary — Q3"  # narrative
    for c, h in enumerate(["KPI", "Target", "Actual"], start=1):
        ws2.cell(row=3, column=c, value=h)
    for r, row_vals in enumerate(
        [("Revenue", 1_000_000, 1_050_000), ("NPS", 70, 74), ("Churn %", 5.0, 4.2)],
        start=4,
    ):
        for c, val in enumerate(row_vals, start=1):
            ws2.cell(row=r, column=c, value=val)

    wb.save(path)


@pytest.fixture(scope="session")
def xlsx_messy(tmp_path_factory: pytest.TempPathFactory) -> Path:
    p = tmp_path_factory.mktemp("excel") / "messy_multi_table.xlsx"
    _build_messy_multi_table(p)
    return p


@pytest.fixture(scope="session")
def xlsx_merged_headers(tmp_path_factory: pytest.TempPathFactory) -> Path:
    p = tmp_path_factory.mktemp("excel") / "merged_headers.xlsx"
    _build_merged_headers(p)
    return p


@pytest.fixture(scope="session")
def xlsx_multi_sheet(tmp_path_factory: pytest.TempPathFactory) -> Path:
    p = tmp_path_factory.mktemp("excel") / "multi_sheet.xlsx"
    _build_multi_sheet(p)
    return p
